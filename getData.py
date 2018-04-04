#!/usr/bin/env python3
# -*- coding: UTF-8 -*-
import openpyxl
import datetime
from urllib import request
from bs4 import BeautifulSoup
from selenium import webdriver

# http://openpyxl.readthedocs.io/en/stable/usage.html

FXH = 'https://www.feixiaohao.com/'
ALC = 'https://www.aicoin.net.cn/currencies'
BLC = 'https://block.cc/'
BLC_API = 'https://data.block.cc/api/v1'
coin = {}
FILENAME = '虚拟币爬取.xlsx'

## 获取时间
TODAY = datetime.date.today()

#新建excel并写入表头
def creatwb(wbname):
    wb=openpyxl.Workbook()
    ws = wb.create_sheet("sheet", 0)
    titleZh = ['序号', '代码', '名称', TODAY]
    for i in range(len(titleZh)):
        ws.cell(row=1, column=i + 1, value=str(titleZh[i]))
    wb.save(filename=wbname)
    print ("新建Excel："+wbname+"成功")

# 判断文件是否存在
def fileExist(wbname):
    wb = openpyxl.load_workbook(wbname)
    sheet = wb.active
    if wb.active is None:
        creatwb(wbname)
    else:
        # 判断表是否有数据
        print('写入当前日期到表头%s'%(sheet.max_column + 1))
        lastdayColunm = sheet.cell(row=1, column=sheet.max_column)
        LASTDAY = lastdayColunm.value
        print('上一次写入的日期是%s'%LASTDAY)
        if (sheet.max_row > 1 and sheet.max_column > 1 and LASTDAY is not TODAY):
            sheet.cell(row=1, column=sheet.max_column + 1, value=TODAY)
            wb.save(wbname)
# 写入表头
def writeTitle(sheet, data):
    pass

# 写入数据
def write07Excel(index, data):
    wb = openpyxl.load_workbook(FILENAME)
    sheet = wb.active
    sheet.title = 'sheet'

    # 从第二行开始写入数据
    for j in range(len(data)):
        sheet.cell(row=index + 2, column=j+1, value=str(list(data.values())[j]))

    wb.save(FILENAME)
    print("写入数据成功！")

# 判断将要写入的数据与表格内的数据的code是否一致
def isEqual():
    pass

# 爬取数据
def getNetData():
    head = {}
    # 写入User Agent信息
    head[
        'User-Agent'] = 'Mozilla/5.0 (Linux; Android 4.1.1; Nexus 7 Build/JRO03D) AppleWebKit/535.19 (KHTML, like Gecko) Chrome/18.0.1025.166  Safari/535.19'
    # 创建Request对象
    req = request.Request(FXH, headers=head)
    # 传入创建好的Request对象
    response = request.urlopen(req)
    # 读取响应信息并解码
    html = response.read().decode('utf-8')
    # 获取数据
    soup_texts = BeautifulSoup(html, 'lxml')
    texts = soup_texts.find(id='table')
    trs = texts.tbody.findAll('tr')
    # 读取数据
    for index in range(len(trs)):
        tds = trs[index].findAll('td')
        coin.clear()
        coin.setdefault('index', tds[0].string)
        coin.setdefault('code', trs[index]['id'])
        coin.setdefault('name', tds[1].a.img['alt'])
        coin.setdefault('price', tds[3].a['data-cny'])
        # coin.setdefault('usPrice', tds[3].a['data-usd'])
        write07Excel(index, coin)

if __name__ == "__main__":

    # 判断表格是否存在
    fileExist(FILENAME)
    getNetData()






